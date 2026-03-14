"""
AetherForge Analytics Module
============================
Grounded local-file analysis for PDFs, markdown/text files, CSVs, and Excel
workbooks. The module generates in-chat artifacts (reports and visuals) that are
served from ``data/generated`` and rendered directly by the frontend.
"""

from __future__ import annotations

import csv
import json
import os
import re
import time
from collections import Counter
from pathlib import Path
from statistics import mean
from typing import Any

import structlog

from src.modules.base import BaseModule

logger = structlog.get_logger("aetherforge.modules.analytics")

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency in some envs
    load_workbook = None

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover - optional dependency in some envs
    PdfReader = None


_SUPPORTED_TABLE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}
_SUPPORTED_TEXT_EXTENSIONS = {".pdf", ".md", ".txt", ".json"}
_SUPPORTED_FILE_EXTENSIONS = _SUPPORTED_TABLE_EXTENSIONS | _SUPPORTED_TEXT_EXTENSIONS
_STOPWORDS = {
    "about",
    "after",
    "also",
    "another",
    "been",
    "being",
    "between",
    "could",
    "does",
    "from",
    "have",
    "into",
    "just",
    "like",
    "more",
    "most",
    "over",
    "same",
    "should",
    "such",
    "than",
    "that",
    "their",
    "them",
    "then",
    "there",
    "these",
    "they",
    "this",
    "through",
    "under",
    "what",
    "when",
    "where",
    "which",
    "while",
    "with",
    "would",
}


def _slugify(value: str, fallback: str = "artifact") -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug[:48] or fallback


def _clean_text(value: str) -> str:
    return " ".join((value or "").split())


def _coerce_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _dedupe(items: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for item in items:
        cleaned = item.strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            deduped.append(cleaned)
    return deduped


class AnalyticsModule(BaseModule):
    """Grounded local analytics and artifact generation."""

    def __init__(self) -> None:
        super().__init__(name="analytics")

    @property
    def system_prompt_extension(self) -> str:
        return (
            "\n\nYou are in Analytics mode. Open local files, ground every claim in the file "
            "contents, generate charts or diagrams when useful, and produce downloadable "
            "Markdown/PDF artifacts when the user asks for exports."
        )

    async def initialize(self) -> None:
        pass

    async def process(self, payload: dict[str, Any], state: Any = None) -> dict[str, Any]:
        return {
            "content": "[AnalyticsModule] Local analysis engine ready.",
            "metadata": {"generated_dir": str(self._generated_dir(state))},
            "causal_edges": [
                {"source": "analytics_start", "target": "ready", "label": "Engine Ready"}
            ],
        }

    def get_tool_definitions(self) -> list[dict[str, Any]]:
        return [
            {
                "name": "analyze_data",
                "description": (
                    "Open a local PDF, Markdown/text file, CSV, or Excel workbook from the "
                    "LiveFolder/uploads area. Ground the explanation in the file contents, "
                    "optionally generate visuals, and optionally export a Markdown or PDF report."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source": {
                            "type": "string",
                            "description": "File name or path, for example 'sales.xlsx' or 'report.pdf'.",
                        },
                        "question": {
                            "type": "string",
                            "description": "The user's actual question or requested framing.",
                        },
                        "report_title": {
                            "type": "string",
                            "description": "Optional title for generated artifacts.",
                        },
                        "format": {
                            "type": "string",
                            "description": "Export format: 'markdown', 'pdf', 'both', or 'csv'.",
                        },
                        "audience": {
                            "type": "string",
                            "description": "Target audience such as 'general' or '10-year-old'.",
                        },
                        "include_visual": {
                            "type": "boolean",
                            "description": "Whether to generate a chart, diagram, or flowchart attachment.",
                        },
                        "visual_type": {
                            "type": "string",
                            "description": "Preferred visual: 'auto', 'bar', 'line', 'pie', 'scatter', or 'flowchart'.",
                        },
                        "x_column": {
                            "type": "string",
                            "description": "Preferred x-axis/category column for tabular files.",
                        },
                        "y_column": {
                            "type": "string",
                            "description": "Preferred numeric column for tabular files.",
                        },
                    },
                    "required": ["source"],
                },
            },
            {
                "name": "create_visual",
                "description": (
                    "Generate a PNG chart or diagram. You can either provide explicit labels/values "
                    "or point to a local file and let the tool derive the visual automatically."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source": {
                            "type": "string",
                            "description": "Optional local file name or path to visualize.",
                        },
                        "chart_type": {
                            "type": "string",
                            "description": (
                                "Type of chart: 'bar', 'line', 'pie', 'scatter', 'hbar', "
                                "'area', 'table', or 'flowchart'."
                            ),
                        },
                        "title": {"type": "string", "description": "Title of the visual."},
                        "labels": {
                            "type": "array",
                            "description": "Labels or category names.",
                            "items": {"type": "string"},
                        },
                        "values": {
                            "type": "array",
                            "description": "Numeric values aligned to the labels.",
                            "items": {"type": "number"},
                        },
                        "steps": {
                            "type": "array",
                            "description": "Optional ordered steps for a flowchart.",
                            "items": {"type": "string"},
                        },
                        "x_label": {"type": "string", "description": "Optional x-axis label."},
                        "y_label": {"type": "string", "description": "Optional y-axis label."},
                        "x_column": {
                            "type": "string",
                            "description": "Preferred x-axis/category column when 'source' is used.",
                        },
                        "y_column": {
                            "type": "string",
                            "description": "Preferred numeric column when 'source' is used.",
                        },
                    },
                    "required": ["chart_type", "title"],
                },
            },
        ]

    def execute_tool(self, name: str, args: dict[str, Any], state: Any | None = None) -> Any:
        if name == "create_visual":
            return self._handle_create_visual(args, state)
        if name == "analyze_data":
            return self._handle_analyze_data(args, state)
        return {"content": f"Error: Tool '{name}' not found in Analytics module."}

    def _handle_analyze_data(self, args: dict[str, Any], state: Any | None = None) -> dict[str, Any]:
        source = str(args.get("source", "")).strip()
        if not source:
            return {"content": "Analytics could not run because no source file was provided."}

        path = self._resolve_source_path(source, state)
        if path is None:
            return {
                "content": (
                    f"Analytics could not find '{source}' in the local data folders. "
                    "Try a file from data/LiveFolder or data/uploads."
                )
            }

        fmt = str(args.get("format", "markdown")).lower()
        report_title = str(args.get("report_title") or f"Analysis of {path.name}")
        question = str(args.get("question", "")).strip()
        audience = str(args.get("audience", "general")).strip() or "general"
        include_visual = bool(args.get("include_visual", False))
        visual_type = str(args.get("visual_type", "auto")).strip().lower() or "auto"

        if path.suffix.lower() in _SUPPORTED_TABLE_EXTENSIONS:
            analysis = self._analyze_tabular_file(path, args)
        elif path.suffix.lower() in _SUPPORTED_TEXT_EXTENSIONS:
            analysis = self._analyze_document_file(path, state)
        else:
            return {
                "content": (
                    f"Analytics found '{path.name}' but does not support '{path.suffix}' yet. "
                    "Use PDF, Markdown/text, CSV, or Excel."
                )
            }

        attachments: list[str] = []
        image_paths: list[Path] = []

        if include_visual:
            visual_paths = self._build_visuals_for_analysis(
                analysis=analysis,
                path=path,
                state=state,
                preferred_visual=visual_type,
                x_column=str(args.get("x_column", "")),
                y_column=str(args.get("y_column", "")),
            )
            image_paths.extend(visual_paths)
            attachments.extend(visual_path.name for visual_path in visual_paths)

        report_markdown = self._build_report_markdown(
            title=report_title,
            path=path,
            question=question,
            audience=audience,
            analysis=analysis,
            image_paths=image_paths,
        )

        generated_dir = self._generated_dir(state)
        report_slug = _slugify(report_title, "analysis_report")
        ts = int(time.time())

        if fmt in {"markdown", "md", "both"}:
            md_filename = f"report_{report_slug}_{ts}.md"
            (generated_dir / md_filename).write_text(report_markdown, encoding="utf-8")
            attachments.append(md_filename)

        if fmt in {"pdf", "both"}:
            pdf_filename = f"report_{report_slug}_{ts}.pdf"
            pdf_path = generated_dir / pdf_filename
            pdf_bytes = self._markdown_to_pdf(report_markdown, state)
            if pdf_bytes is not None:
                pdf_path.write_bytes(pdf_bytes)
                attachments.append(pdf_filename)

        if fmt == "csv" and analysis.get("headers") and analysis.get("rows"):
            csv_filename = f"report_{report_slug}_{ts}.csv"
            self._write_csv(generated_dir / csv_filename, analysis["headers"], analysis["rows"])
            attachments.append(csv_filename)

        attachments = _dedupe(attachments)
        content_lines = [
            f"I inspected `{path.name}` and grounded the output in the local file contents.",
            "",
            "Grounded notes:",
        ]
        content_lines.extend(f"- {point}" for point in analysis["summary_points"])

        if question:
            content_lines.append("")
            content_lines.append(f"Requested framing: {question}")
        if audience:
            content_lines.append(f"Target audience: {audience}")

        if analysis["citations"]:
            content_lines.append("")
            content_lines.append("Source references:")
            for citation in analysis["citations"][:6]:
                locator = citation["source"]
                if citation.get("page") not in (None, ""):
                    locator += f" | p.{citation['page']}"
                if citation.get("section"):
                    locator += f" | {citation['section']}"
                snippet = citation.get("snippet")
                if snippet:
                    content_lines.append(f"{citation['label']} {locator} - {snippet}")
                else:
                    content_lines.append(f"{citation['label']} {locator}")

        if attachments:
            content_lines.append("")
            content_lines.append("Generated artifacts:")
            content_lines.extend(f"[attachment:{name}]" for name in attachments)

        return {
            "content": "\n".join(content_lines).strip(),
            "attachments": attachments,
            "citations": analysis["citations"],
        }

    def _handle_create_visual(
        self,
        args: dict[str, Any],
        state: Any | None = None,
    ) -> dict[str, Any]:
        source = str(args.get("source", "")).strip()
        if source:
            source_path = self._resolve_source_path(source, state)
            if source_path is None:
                return {
                    "content": (
                        f"Visual generation could not find '{source}' in the local data folders."
                    )
                }

            if source_path.suffix.lower() in _SUPPORTED_TABLE_EXTENSIONS:
                analysis = self._analyze_tabular_file(source_path, args)
            else:
                analysis = self._analyze_document_file(source_path, state)

            visual_paths = self._build_visuals_for_analysis(
                analysis=analysis,
                path=source_path,
                state=state,
                preferred_visual=str(args.get("chart_type", "auto")).lower(),
                x_column=str(args.get("x_column", "")),
                y_column=str(args.get("y_column", "")),
            )
            if not visual_paths:
                return {"content": f"Could not derive a visual from '{source_path.name}'."}

            attachments = [visual_path.name for visual_path in visual_paths]
            return {
                "content": "\n".join(
                    [
                        f"Generated visual artifact(s) for `{source_path.name}`.",
                        "",
                        *[f"[attachment:{name}]" for name in attachments],
                    ]
                ),
                "attachments": attachments,
                "citations": analysis["citations"],
            }

        chart_type = str(args.get("chart_type", "bar")).lower()
        title = str(args.get("title", "Chart")).strip() or "Chart"
        labels = [str(item) for item in args.get("labels", [])]
        values = [value for value in args.get("values", [])]
        steps = [str(item) for item in args.get("steps", [])]

        if chart_type == "flowchart" and steps and not labels:
            labels = steps
        if chart_type == "flowchart" and labels and not values:
            values = list(range(1, len(labels) + 1))

        if chart_type != "flowchart" and (not labels or not values):
            return {"content": "Visual generation needs labels and values when no source file is provided."}

        output_path = self._render_visual(
            chart_type=chart_type,
            title=title,
            labels=labels,
            values=values,
            state=state,
            x_label=str(args.get("x_label", "")),
            y_label=str(args.get("y_label", "")),
        )
        return {
            "content": f"Generated `{title}`.\n\n[attachment:{output_path.name}]",
            "attachments": [output_path.name],
            "citations": [],
        }

    def _analyze_tabular_file(self, path: Path, args: dict[str, Any]) -> dict[str, Any]:
        headers, rows = self._load_tabular_rows(path)
        non_empty_rows = [row for row in rows if any(str(cell).strip() for cell in row)]
        sample_rows = non_empty_rows[:6]
        row_count = len(non_empty_rows)
        column_count = len(headers)

        numeric_columns: list[dict[str, Any]] = []
        categorical_columns: list[dict[str, Any]] = []
        for idx, header in enumerate(headers):
            values = [row[idx] if idx < len(row) else "" for row in non_empty_rows]
            non_empty = [value for value in values if str(value).strip()]
            numeric_values = [_coerce_number(value) for value in non_empty]
            parsed_numbers = [value for value in numeric_values if value is not None]
            if non_empty and parsed_numbers and len(parsed_numbers) >= max(2, int(len(non_empty) * 0.6)):
                numeric_columns.append(
                    {
                        "index": idx,
                        "name": header,
                        "values": parsed_numbers,
                    }
                )
            else:
                categorical_columns.append(
                    {
                        "index": idx,
                        "name": header,
                        "values": [str(value).strip() for value in non_empty if str(value).strip()],
                    }
                )

        summary_points = [
            f"Rows analysed: {row_count}.",
            f"Columns: {column_count} ({', '.join(headers[:8])}).",
        ]
        for column in numeric_columns[:3]:
            values = column["values"]
            summary_points.append(
                (
                    f"Numeric column '{column['name']}' ranges from {min(values):.2f} to "
                    f"{max(values):.2f} with an average of {mean(values):.2f}."
                )
            )
        for column in categorical_columns[:2]:
            if not column["values"]:
                continue
            common = Counter(column["values"]).most_common(3)
            summary_points.append(
                f"Common values in '{column['name']}': "
                + ", ".join(f"{label} ({count})" for label, count in common)
                + "."
            )

        citations = [
            {
                "label": "[1]",
                "source": path.name,
                "page": None,
                "section": "Tabular summary",
                "snippet": f"{row_count} rows x {column_count} columns.",
                "kind": "table",
            }
        ]

        return {
            "kind": "table",
            "headers": headers,
            "rows": sample_rows,
            "all_rows": non_empty_rows,
            "summary_points": summary_points,
            "citations": citations,
            "numeric_columns": numeric_columns,
            "categorical_columns": categorical_columns,
        }

    def _analyze_document_file(self, path: Path, state: Any | None = None) -> dict[str, Any]:
        citations: list[dict[str, Any]] = []
        sections: list[str] = []
        raw_text_blocks: list[str] = []

        sparse_index = getattr(state, "sparse_index", None)
        if sparse_index is not None and hasattr(sparse_index, "get_chunks_by_source"):
            try:
                docs = sparse_index.get_chunks_by_source(path.name, limit=18)
            except Exception as exc:  # pragma: no cover - defensive logging
                logger.warning("Failed to read indexed chunks", source=path.name, error=str(exc))
                docs = []
            for idx, doc in enumerate(docs[:6], start=1):
                meta = getattr(doc, "metadata", {}) or {}
                section = str(meta.get("section", "")).strip()
                if section and section.lower() != "unknown":
                    sections.append(section)
                snippet = _clean_text(str(getattr(doc, "page_content", "")))
                if snippet:
                    raw_text_blocks.append(snippet)
                    citations.append(
                        {
                            "label": f"[{idx}]",
                            "source": str(meta.get("source", path.name)),
                            "page": meta.get("page"),
                            "section": section or None,
                            "snippet": snippet[:240],
                            "kind": "document",
                        }
                    )

        if not raw_text_blocks:
            raw_text, direct_citations, direct_sections = self._extract_document_text(path)
            raw_text_blocks = [raw_text] if raw_text else []
            citations = direct_citations
            sections = direct_sections

        combined_text = "\n\n".join(block for block in raw_text_blocks if block)
        keywords = self._top_keywords(combined_text)
        if not sections:
            sections = self._extract_headings(combined_text)

        summary_points = [f"Document type: {path.suffix.lower().lstrip('.').upper()}."]
        if sections:
            summary_points.append(f"Likely main sections: {', '.join(sections[:6])}.")
        if keywords:
            summary_points.append(f"Frequent concepts: {', '.join(keywords[:6])}.")
        if citations:
            summary_points.append(f"Grounded in {len(citations)} cited excerpt(s) from the local file.")

        return {
            "kind": "document",
            "summary_points": summary_points,
            "citations": citations,
            "sections": sections,
            "keywords": keywords,
            "combined_text": combined_text,
        }

    def _build_visuals_for_analysis(
        self,
        *,
        analysis: dict[str, Any],
        path: Path,
        state: Any | None,
        preferred_visual: str,
        x_column: str,
        y_column: str,
    ) -> list[Path]:
        visuals: list[Path] = []
        preferred = preferred_visual or "auto"

        if analysis["kind"] == "table":
            spec = self._choose_tabular_visual(
                headers=analysis["headers"],
                rows=analysis["all_rows"],
                source_name=path.name,
                preferred_visual=preferred,
                x_column=x_column,
                y_column=y_column,
            )
            if spec is not None:
                visuals.append(self._render_visual(state=state, **spec))
            return visuals

        steps = analysis.get("sections") or self._derive_steps_from_text(analysis.get("combined_text", ""))
        if preferred in {"auto", "flowchart"} and len(steps) >= 2:
            visuals.append(
                self._render_visual(
                    chart_type="flowchart",
                    title=f"Flow of {path.name}",
                    labels=steps[:6],
                    values=list(range(1, len(steps[:6]) + 1)),
                    state=state,
                )
            )

        keywords = analysis.get("keywords", [])
        if keywords:
            keyword_counts = Counter(
                word
                for word in re.findall(r"[A-Za-z][A-Za-z-]{3,}", analysis.get("combined_text", "").lower())
                if word not in _STOPWORDS
            )
            labels = []
            values = []
            for keyword, count in keyword_counts.most_common(6):
                labels.append(keyword)
                values.append(count)
            if labels and values:
                visuals.append(
                    self._render_visual(
                        chart_type="bar" if preferred == "auto" else preferred,
                        title=f"Key concepts in {path.name}",
                        labels=labels,
                        values=values,
                        state=state,
                        x_label="Concept",
                        y_label="Mentions",
                    )
                )

        return visuals[:2]

    def _choose_tabular_visual(
        self,
        *,
        headers: list[str],
        rows: list[list[str]],
        source_name: str,
        preferred_visual: str,
        x_column: str,
        y_column: str,
    ) -> dict[str, Any] | None:
        header_index = {header.lower(): idx for idx, header in enumerate(headers)}
        x_idx = header_index.get(x_column.strip().lower()) if x_column.strip() else None
        y_idx = header_index.get(y_column.strip().lower()) if y_column.strip() else None

        numeric_candidates: list[tuple[int, str]] = []
        categorical_candidates: list[tuple[int, str]] = []
        for idx, header in enumerate(headers):
            values = [row[idx] if idx < len(row) else "" for row in rows]
            non_empty = [value for value in values if str(value).strip()]
            parsed = [_coerce_number(value) for value in non_empty]
            numbers = [value for value in parsed if value is not None]
            if non_empty and numbers and len(numbers) >= max(2, int(len(non_empty) * 0.6)):
                numeric_candidates.append((idx, header))
            else:
                categorical_candidates.append((idx, header))

        if y_idx is None and numeric_candidates:
            y_idx = numeric_candidates[0][0]
        if x_idx is None and categorical_candidates:
            x_idx = categorical_candidates[0][0]

        labels: list[str] = []
        values: list[float] = []
        x_label = headers[x_idx] if x_idx is not None else "Item"
        y_label = headers[y_idx] if y_idx is not None else "Value"

        if x_idx is not None and y_idx is not None:
            grouped: dict[str, float] = {}
            for row in rows:
                if y_idx >= len(row):
                    continue
                label = str(row[x_idx]).strip() if x_idx < len(row) else ""
                number = _coerce_number(row[y_idx])
                if not label or number is None:
                    continue
                grouped[label] = grouped.get(label, 0.0) + number
            for label, value in sorted(grouped.items(), key=lambda item: abs(item[1]), reverse=True)[:8]:
                labels.append(label)
                values.append(value)

        if not labels and y_idx is not None:
            for idx, row in enumerate(rows[:8], start=1):
                if y_idx >= len(row):
                    continue
                number = _coerce_number(row[y_idx])
                if number is None:
                    continue
                labels.append(f"Row {idx}")
                values.append(number)

        if not labels or not values:
            return None

        chart_type = preferred_visual if preferred_visual not in {"", "auto"} else "bar"
        if chart_type == "flowchart":
            chart_type = "bar"

        return {
            "chart_type": chart_type,
            "title": f"Visual summary of {source_name}",
            "labels": labels,
            "values": values,
            "x_label": x_label,
            "y_label": y_label,
        }

    def _load_tabular_rows(self, path: Path) -> tuple[list[str], list[list[str]]]:
        if path.suffix.lower() in {".csv", ".tsv"}:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                all_rows = [list(row) for row in reader]
        else:
            if load_workbook is None:
                raise RuntimeError("openpyxl is required for Excel support.")
            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            all_rows = [
                ["" if cell is None else str(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            workbook.close()

        if not all_rows:
            return [], []

        headers = [str(cell).strip() or f"Column {idx + 1}" for idx, cell in enumerate(all_rows[0])]
        normalized_rows: list[list[str]] = []
        for row in all_rows[1:]:
            padded = list(row) + [""] * max(0, len(headers) - len(row))
            normalized_rows.append([str(cell).strip() for cell in padded[: len(headers)]])
        return headers, normalized_rows

    def _extract_document_text(self, path: Path) -> tuple[str, list[dict[str, Any]], list[str]]:
        if path.suffix.lower() == ".pdf":
            if PdfReader is None:
                raise RuntimeError("pypdf is required for PDF analysis.")
            reader = PdfReader(str(path))
            sections: list[str] = []
            citations: list[dict[str, Any]] = []
            page_texts: list[str] = []
            for page_number, page in enumerate(reader.pages[:8], start=1):
                text = _clean_text(page.extract_text() or "")
                if not text:
                    continue
                page_texts.append(text)
                sections.extend(self._extract_headings(text))
                citations.append(
                    {
                        "label": f"[{len(citations) + 1}]",
                        "source": path.name,
                        "page": page_number,
                        "section": None,
                        "snippet": text[:240],
                        "kind": "document",
                    }
                )
            return "\n\n".join(page_texts), citations, _dedupe(sections)

        text = path.read_text(encoding="utf-8", errors="ignore")
        clean = _clean_text(text)
        sections = self._extract_headings(text)
        citations = [
            {
                "label": "[1]",
                "source": path.name,
                "page": None,
                "section": sections[0] if sections else None,
                "snippet": clean[:240],
                "kind": "document",
            }
        ]
        return clean, citations, sections

    def _extract_headings(self, text: str) -> list[str]:
        headings: list[str] = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith("#"):
                headings.append(line.lstrip("# ").strip())
                continue
            if 4 <= len(line) <= 80 and len(line.split()) <= 8:
                if line.isupper():
                    headings.append(line.title())
                    continue
                if line == line.title():
                    headings.append(line)
        return _dedupe(headings)

    def _derive_steps_from_text(self, text: str) -> list[str]:
        paragraphs = [
            _clean_text(chunk)
            for chunk in re.split(r"\n\s*\n", text)
            if _clean_text(chunk)
        ]
        steps: list[str] = []
        for paragraph in paragraphs[:6]:
            sentence = re.split(r"(?<=[.!?])\s+", paragraph)[0]
            if len(sentence) > 96:
                sentence = sentence[:93].rstrip() + "..."
            steps.append(sentence)
        return _dedupe(steps)

    def _top_keywords(self, text: str, limit: int = 6) -> list[str]:
        counts = Counter(
            word
            for word in re.findall(r"[A-Za-z][A-Za-z-]{3,}", text.lower())
            if word not in _STOPWORDS
        )
        return [word for word, _count in counts.most_common(limit)]

    def _build_report_markdown(
        self,
        *,
        title: str,
        path: Path,
        question: str,
        audience: str,
        analysis: dict[str, Any],
        image_paths: list[Path],
    ) -> str:
        lines = [
            f"# {title}",
            "",
            f"**Source file:** `{path.name}`",
            f"**Audience:** {audience}",
        ]
        if question:
            lines.extend(["", "## Requested Task", "", question])

        lines.extend(["", "## Grounded Notes", ""])
        lines.extend(f"- {point}" for point in analysis["summary_points"])

        if analysis["kind"] == "table" and analysis.get("headers") and analysis.get("rows"):
            lines.extend(["", "## Sample Rows", ""])
            headers = analysis["headers"]
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join("---" for _ in headers) + " |")
            for row in analysis["rows"]:
                lines.append("| " + " | ".join(str(cell) for cell in row) + " |")

        if analysis["kind"] == "document" and analysis.get("sections"):
            lines.extend(["", "## Document Flow", ""])
            lines.extend(f"1. {step}" for step in analysis["sections"][:6])

        if image_paths:
            lines.extend(["", "## Generated Visuals", ""])
            for image_path in image_paths:
                lines.append(f"![{image_path.name}](file://{image_path.resolve()})")

        if analysis["citations"]:
            lines.extend(["", "## Sources", ""])
            for citation in analysis["citations"]:
                locator = citation["source"]
                if citation.get("page") not in (None, ""):
                    locator += f" | p.{citation['page']}"
                if citation.get("section"):
                    locator += f" | {citation['section']}"
                snippet = citation.get("snippet")
                if snippet:
                    lines.append(f"- {citation['label']} {locator}: {snippet}")
                else:
                    lines.append(f"- {citation['label']} {locator}")

        lines.extend(
            [
                "",
                "---",
                f"*Generated by AetherForge Analytics at {time.strftime('%Y-%m-%d %H:%M:%S')}*",
            ]
        )
        return "\n".join(lines)

    def _markdown_to_pdf(self, markdown_text: str, state: Any | None) -> bytes | None:
        export_engine = getattr(state, "export_engine", None)
        if export_engine is None or not hasattr(export_engine, "_markdown_to_pdf"):
            return None
        try:
            return export_engine._markdown_to_pdf(markdown_text)
        except Exception as exc:  # pragma: no cover - defensive logging
            logger.warning("PDF export failed", error=str(exc))
            return None

    def _write_csv(self, path: Path, headers: list[str], rows: list[list[str]]) -> None:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)

    def _generated_dir(self, state: Any | None = None) -> Path:
        settings = getattr(state, "settings", None)
        if settings is not None and getattr(settings, "generated_dir", None) is not None:
            generated_dir = Path(settings.generated_dir)
        else:
            generated_dir = Path(os.environ.get("DATA_DIR", "data")) / "generated"
        generated_dir.mkdir(parents=True, exist_ok=True)
        return generated_dir

    def _resolve_source_path(self, source: str, state: Any | None = None) -> Path | None:
        candidate = Path(source).expanduser()
        if candidate.is_absolute() and candidate.exists():
            return candidate

        settings = getattr(state, "settings", None)
        roots: list[Path] = []
        if settings is not None:
            roots.extend(
                [
                    Path(settings.live_folder),
                    Path(settings.uploads_dir),
                    Path(settings.data_dir),
                    Path(settings.generated_dir),
                ]
            )
        else:
            base = Path(os.environ.get("DATA_DIR", "data"))
            roots.extend([base / "LiveFolder", base / "uploads", base, base / "generated"])

        source_name = Path(source).name
        for root in roots:
            direct_match = root / source
            if direct_match.exists():
                return direct_match
            by_name_match = root / source_name
            if by_name_match.exists():
                return by_name_match
        return None

    def _render_visual(
        self,
        *,
        chart_type: str,
        title: str,
        labels: list[str],
        values: list[Any],
        state: Any | None,
        x_label: str = "",
        y_label: str = "",
    ) -> Path:
        try:
            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.patches as mpatches
            import matplotlib.pyplot as plt
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise RuntimeError("matplotlib is required for chart generation.") from exc

        generated_dir = self._generated_dir(state)
        chart_type = chart_type.lower()
        numeric_values = [_coerce_number(value) or 0.0 for value in values]
        filename = f"chart_{_slugify(title, 'chart')}_{int(time.time() * 1000)}.png"
        filepath = generated_dir / filename

        plt.style.use("dark_background")
        fig, ax = plt.subplots(figsize=(10, 6))
        colors = [
            "#2563eb",
            "#14b8a6",
            "#f97316",
            "#e11d48",
            "#84cc16",
            "#a855f7",
            "#06b6d4",
            "#f59e0b",
        ]

        if chart_type == "bar":
            bars = ax.bar(labels, numeric_values, color=colors[: len(labels)], alpha=0.9)
            for bar, value in zip(bars, numeric_values):
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    f"{value:.1f}",
                    ha="center",
                    va="bottom",
                    fontsize=9,
                    color="white",
                )
        elif chart_type == "hbar":
            ax.barh(labels, numeric_values, color=colors[: len(labels)], alpha=0.9)
        elif chart_type == "line":
            ax.plot(labels, numeric_values, marker="o", linewidth=2.2, color=colors[0])
            ax.fill_between(range(len(numeric_values)), numeric_values, alpha=0.15, color=colors[0])
        elif chart_type == "scatter":
            ax.scatter(range(len(numeric_values)), numeric_values, c=colors[: len(labels)], s=120)
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=45, ha="right")
        elif chart_type == "pie":
            ax.pie(
                numeric_values,
                labels=labels,
                autopct="%1.1f%%",
                startangle=90,
                colors=colors[: len(labels)],
                textprops={"color": "white"},
            )
        elif chart_type == "area":
            ax.fill_between(range(len(numeric_values)), numeric_values, alpha=0.4, color=colors[0])
            ax.plot(range(len(numeric_values)), numeric_values, color=colors[0], linewidth=2)
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=45, ha="right")
        elif chart_type == "flowchart":
            ax.set_xlim(0, 10)
            ax.set_ylim(0, len(labels) * 2 + 1)
            ax.axis("off")
            for idx, label in enumerate(labels):
                y_pos = len(labels) * 2 - idx * 2
                box = mpatches.FancyBboxPatch(
                    (1.6, y_pos - 0.45),
                    6.8,
                    0.9,
                    boxstyle="round,pad=0.28",
                    facecolor=colors[idx % len(colors)],
                    edgecolor="white",
                    linewidth=1.2,
                )
                ax.add_patch(box)
                ax.text(5, y_pos, label, ha="center", va="center", fontsize=10.5, color="white")
                if idx < len(labels) - 1:
                    ax.annotate(
                        "",
                        xy=(5, y_pos - 0.55),
                        xytext=(5, y_pos - 1.45),
                        arrowprops=dict(arrowstyle="->", color="white", lw=2),
                    )
        elif chart_type == "table":
            ax.axis("off")
            table_data = [[label, f"{value:.2f}"] for label, value in zip(labels, numeric_values)]
            table = ax.table(
                cellText=table_data,
                colLabels=["Item", "Value"],
                cellLoc="center",
                loc="center",
            )
            table.auto_set_font_size(False)
            table.set_fontsize(11)
            table.scale(1.2, 1.5)
        else:
            ax.bar(labels, numeric_values, color=colors[: len(labels)], alpha=0.9)

        ax.set_title(title, fontsize=16, fontweight="bold", pad=16, color="white")
        if chart_type not in {"pie", "flowchart", "table"}:
            if x_label:
                ax.set_xlabel(x_label)
            if y_label:
                ax.set_ylabel(y_label)
            ax.tick_params(axis="x", rotation=45)
            ax.grid(axis="y", alpha=0.2)

        plt.tight_layout()
        fig.savefig(filepath, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)
        logger.info("Generated visual artifact", file=str(filepath), chart_type=chart_type)
        return filepath
